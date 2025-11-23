"""
Export utilities for Excel and PDF generation.

Provides functions to export ReportData to Excel and PDF formats.
"""
import logging
from typing import List, Dict
from datetime import datetime
from dataclasses import asdict

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export data to Excel format."""

    @staticmethod
    def export_report(report_data, file_path: str):
        """
        Export report to Excel file.

        Args:
            report_data: ReportData instance
            file_path: Output file path

        Requires:
            openpyxl: pip install openpyxl
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            # Fallback to CSV
            ExcelExporter._export_to_csv(report_data, file_path.replace('.xlsx', '.csv'))
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "보고서"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header info
        ws['A1'] = "보고서 종류"
        ws['B1'] = report_data.report_type
        ws['A2'] = "생성 시간"
        ws['B2'] = datetime.fromisoformat(report_data.generated_at).strftime("%Y-%m-%d %H:%M:%S")

        if report_data.date_range_start:
            ws['A3'] = "기간"
            if report_data.date_range_end and report_data.date_range_end != report_data.date_range_start:
                ws['B3'] = f"{report_data.date_range_start} ~ {report_data.date_range_end}"
            else:
                ws['B3'] = report_data.date_range_start

        # Write summary
        row = 5
        ws[f'A{row}'] = "요약"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1

        if report_data.summary:
            for key, value in report_data.summary.items():
                if not isinstance(value, (dict, list)):
                    ws[f'A{row}'] = key
                    ws[f'B{row}'] = value
                    row += 1

        # Write data table
        if report_data.data:
            row += 2
            ws[f'A{row}'] = "상세 데이터"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1

            # Determine columns
            if report_data.data:
                columns = list(report_data.data[0].keys())

                # Write column headers
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                row += 1

                # Write data rows
                for data_row in report_data.data:
                    for col_idx, col_name in enumerate(columns, start=1):
                        value = data_row.get(col_name, "")
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = border

                    row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save workbook
        wb.save(file_path)
        logger.info(f"Exported report to Excel: {file_path}")

    @staticmethod
    def _export_to_csv(report_data, file_path: str):
        """Fallback CSV export if openpyxl not available."""
        import csv

        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # Write header
            writer.writerow(["보고서 종류", report_data.report_type])
            writer.writerow(["생성 시간", report_data.generated_at])
            writer.writerow([])

            # Write summary
            writer.writerow(["요약"])
            if report_data.summary:
                for key, value in report_data.summary.items():
                    if not isinstance(value, (dict, list)):
                        writer.writerow([key, value])

            writer.writerow([])

            # Write data
            if report_data.data:
                writer.writerow(["상세 데이터"])
                columns = list(report_data.data[0].keys())
                writer.writerow(columns)

                for row in report_data.data:
                    writer.writerow([row.get(col, "") for col in columns])

        logger.info(f"Exported report to CSV (fallback): {file_path}")


class PDFExporter:
    """Export data to PDF format."""

    @staticmethod
    def export_report(report_data, file_path: str):
        """
        Export report to PDF file.

        Args:
            report_data: ReportData instance
            file_path: Output file path

        Requires:
            reportlab: pip install reportlab
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # Try to register Korean font (if available)
            try:
                # Common Korean font locations
                font_paths = [
                    "C:/Windows/Fonts/malgun.ttf",  # Windows
                    "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",  # Linux
                    "/Library/Fonts/AppleGothic.ttf"  # Mac
                ]
                for font_path in font_paths:
                    try:
                        pdfmetrics.registerFont(TTFont('Korean', font_path))
                        korean_available = True
                        break
                    except:
                        continue
                else:
                    korean_available = False
            except:
                korean_available = False

        except ImportError:
            logger.error("reportlab not installed. Run: pip install reportlab")
            # Fallback to text file
            PDFExporter._export_to_text(report_data, file_path.replace('.pdf', '.txt'))
            return

        # Create PDF
        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        if korean_available:
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName='Korean',
                fontSize=18
            )
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName='Korean',
                fontSize=10
            )
        else:
            title_style = styles['Heading1']
            normal_style = styles['Normal']

        # Title
        elements.append(Paragraph(f"보고서: {report_data.report_type}", title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Header info
        info_data = [
            ["생성 시간", datetime.fromisoformat(report_data.generated_at).strftime("%Y-%m-%d %H:%M:%S")]
        ]

        if report_data.date_range_start:
            if report_data.date_range_end and report_data.date_range_end != report_data.date_range_start:
                info_data.append(["기간", f"{report_data.date_range_start} ~ {report_data.date_range_end}"])
            else:
                info_data.append(["기간", report_data.date_range_start])

        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Korean' if korean_available else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Summary
        elements.append(Paragraph("요약", title_style))
        elements.append(Spacer(1, 0.1 * inch))

        if report_data.summary:
            summary_data = []
            for key, value in report_data.summary.items():
                if not isinstance(value, (dict, list)):
                    summary_data.append([str(key), str(value)])

            if summary_data:
                summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Korean' if korean_available else 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(summary_table)

        elements.append(Spacer(1, 0.3 * inch))

        # Data table (first 50 rows for PDF)
        if report_data.data:
            elements.append(Paragraph("상세 데이터 (최대 50행)", title_style))
            elements.append(Spacer(1, 0.1 * inch))

            columns = list(report_data.data[0].keys())[:8]  # Limit columns for PDF
            data_rows = [columns]

            for row in report_data.data[:50]:  # Limit to 50 rows
                data_rows.append([str(row.get(col, ""))[:30] for col in columns])  # Truncate long text

            data_table = Table(data_rows)
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Korean' if korean_available else 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(data_table)

        # Build PDF
        doc.build(elements)
        logger.info(f"Exported report to PDF: {file_path}")

    @staticmethod
    def _export_to_text(report_data, file_path: str):
        """Fallback text export if reportlab not available."""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(f"보고서 종류: {report_data.report_type}\n")
            f.write(f"생성 시간: {report_data.generated_at}\n")

            if report_data.date_range_start:
                f.write(f"기간: {report_data.date_range_start}")
                if report_data.date_range_end and report_data.date_range_end != report_data.date_range_start:
                    f.write(f" ~ {report_data.date_range_end}")
                f.write("\n")

            f.write("\n=== 요약 ===\n")
            if report_data.summary:
                for key, value in report_data.summary.items():
                    if not isinstance(value, (dict, list)):
                        f.write(f"{key}: {value}\n")

            f.write("\n=== 상세 데이터 ===\n")
            if report_data.data:
                columns = list(report_data.data[0].keys())
                f.write("\t".join(columns) + "\n")

                for row in report_data.data:
                    f.write("\t".join(str(row.get(col, "")) for col in columns) + "\n")

        logger.info(f"Exported report to text (fallback): {file_path}")
